import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─── Paleta ───────────────────────────────────────────────────────────────────
COR_VERDE_ESCURO  = "085041"
COR_VERDE_CLARO   = "E1F5EE"
COR_AMARELO_CLARO = "FAEEDA"
COR_AMARELO_ESC   = "633806"
COR_VERMELHO_CLARO= "FCEBEB"
COR_VERMELHO_ESC  = "791F1F"
COR_AZUL_CLARO    = "E6F1FB"
COR_AZUL_ESC      = "0C447C"
COR_CINZA         = "F1EFE8"
COR_CINZA_ESC     = "444441"
BRANCO            = "FFFFFF"

STATUS_CORES = {
    "Conciliado":     (COR_VERDE_CLARO,   COR_VERDE_ESCURO),
    "Divergente":     (COR_AMARELO_CLARO, COR_AMARELO_ESC),
    "Nao Conciliado": (COR_VERMELHO_CLARO,COR_VERMELHO_ESC),
    "So no ERP":      (COR_AZUL_CLARO,    COR_AZUL_ESC),
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(ws, row, cols, bg=COR_VERDE_ESCURO, fg=BRANCO):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=fg, size=10)
        cell.fill = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value else 0 for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


def _fmt_valor(v):
    if v is None or pd.isna(v):
        return "—"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_data(d):
    if d is None or pd.isna(d):
        return "—"
    try:
        return pd.Timestamp(d).strftime("%d/%m/%Y")
    except Exception:
        return str(d)


# ─── Excel completo ───────────────────────────────────────────────────────────

def exportar_excel(df_resultado: pd.DataFrame, resumo_dict: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _aba_resumo(wb, resumo_dict)
    _aba_status(wb, df_resultado, "Conciliados",     ["Conciliado"])
    _aba_status(wb, df_resultado, "Divergentes",     ["Divergente"])
    _aba_status(wb, df_resultado, "Nao Conciliados", ["Nao Conciliado"])
    _aba_status(wb, df_resultado, "So no Banco",     ["So no Banco"])
    _aba_status(wb, df_resultado, "So no ERP",       ["So no ERP"])
    _aba_completa(wb, df_resultado)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _aba_resumo(wb, res: dict):
    ws = wb.create_sheet("Resumo")
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:D1")
    ws["A1"] = "CONCILIAÇÃO BANCÁRIA — RESUMO EXECUTIVO"
    ws["A1"].font = Font(bold=True, size=14, color=BRANCO)
    ws["A1"].fill = _fill(COR_VERDE_ESCURO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Métricas
    metricas = [
        ("Total lançamentos banco",    res["total_banco"],       COR_CINZA,          COR_CINZA_ESC),
        ("Conciliados",                res["conciliados"],        COR_VERDE_CLARO,    COR_VERDE_ESCURO),
        ("Divergentes",                res["divergentes"],        COR_AMARELO_CLARO,  COR_AMARELO_ESC),
        ("Não conciliados",            res["nao_conciliados"],    COR_VERMELHO_CLARO, COR_VERMELHO_ESC),
        ("So no ERP",                  res["so_erp"],             COR_AZUL_CLARO,     COR_AZUL_ESC),
        ("Taxa de conciliação (%)",    f"{res['taxa_pct']}%",     COR_VERDE_CLARO,    COR_VERDE_ESCURO),
    ]

    ws["A3"] = "Indicador"
    ws["B3"] = "Valor"
    _header_style(ws, 3, 2)

    for i, (label, val, bg, fg) in enumerate(metricas, start=4):
        ws.cell(row=i, column=1, value=label).fill = _fill(bg)
        ws.cell(row=i, column=1).font = Font(color=fg, size=10)
        ws.cell(row=i, column=1).border = _thin_border()
        ws.cell(row=i, column=2, value=val).fill = _fill(bg)
        ws.cell(row=i, column=2).font = Font(bold=True, color=fg, size=10)
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2).border = _thin_border()

    # Valores financeiros
    ws["A11"] = "Resumo Financeiro"
    ws["A11"].font = Font(bold=True, size=11, color=BRANCO)
    ws["A11"].fill = _fill(COR_VERDE_ESCURO)
    ws.merge_cells("A11:B11")

    fin = [
        ("Valor conciliado",      _fmt_valor(res["valor_conciliado"])),
        ("Valor divergente",      _fmt_valor(res["valor_divergente"])),
        ("Valor não conciliado",  _fmt_valor(res["valor_nao_conciliado"])),
    ]
    for i, (label, val) in enumerate(fin, start=12):
        ws.cell(row=i, column=1, value=label).border = _thin_border()
        ws.cell(row=i, column=2, value=val).border = _thin_border()
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def _aba_status(wb, df, nome_aba, status_list):
    filtrado = df[df["status"].isin(status_list)].copy()
    ws = wb.create_sheet(nome_aba)
    ws.sheet_view.showGridLines = False

    cabecalhos = [
        "Data Banco", "Valor Banco", "D/C Banco", "Descr. Banco", "Doc. Banco",
        "Data ERP",   "Valor ERP",   "D/C ERP",   "Descr. ERP",   "Doc. ERP",
        "Status", "Confiança", "Dif. Valor (R$)", "Dif. Dias"
    ]

    bg = STATUS_CORES.get(status_list[0], (COR_CINZA, COR_CINZA_ESC))[0]
    fg = STATUS_CORES.get(status_list[0], (COR_CINZA, COR_CINZA_ESC))[1]

    for c, h in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color=fg, size=10)
        cell.fill = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[1].height = 22

    for r, row in enumerate(filtrado.itertuples(), start=2):
        # Status em português
        status_label = {
            "Conciliado":     "Conciliado",
            "Divergente":     "Divergente",
            "Nao Conciliado": "Não Conciliado",
            "So no Banco":    "Só no Banco",
            "So no ERP":      "Só no ERP",
        }.get(row.status, row.status)
        banco_dc = getattr(row, "banco_dc", "") or ""
        erp_dc   = getattr(row, "erp_dc",   "") or ""
        vals = [
            _fmt_data(row.banco_data),  _fmt_valor(row.banco_valor),
            banco_dc,
            row.banco_descricao or "—", row.banco_documento or "—",
            _fmt_data(row.erp_data),    _fmt_valor(row.erp_valor),
            erp_dc,
            row.erp_descricao or "—",   row.erp_documento or "—",
            status_label, row.confianca,
            _fmt_valor(row.diferenca_valor),
            str(row.diferenca_dias) if row.diferenca_dias is not None else "—",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = _thin_border()
            cell.font = Font(size=9)
            if r % 2 == 0:
                cell.fill = _fill("F7F7F5")

    _auto_width(ws)


def _aba_completa(wb, df):
    _aba_status(wb, df, "Todos os Lançamentos", df["status"].unique().tolist())


# ─── CSV input ERP ────────────────────────────────────────────────────────────

def exportar_csv_erp(df_resultado: pd.DataFrame) -> bytes:
    conc = df_resultado[df_resultado["status"] == "Conciliado"].copy()
    conc = conc[["erp_data", "erp_valor", "erp_descricao", "erp_documento"]].copy()
    conc.columns = ["Data", "Valor", "Descricao", "Documento"]
    conc["Data"] = conc["Data"].apply(_fmt_data)
    conc["Valor"] = conc["Valor"].apply(
        lambda v: f"{v:.2f}".replace(".", ",") if pd.notna(v) else ""
    )
    buf = io.StringIO()
    conc.to_csv(buf, index=False, sep=";", encoding="utf-8-sig")
    return buf.getvalue().encode("utf-8-sig")
